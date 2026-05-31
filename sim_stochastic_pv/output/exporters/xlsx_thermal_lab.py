"""
Excel exporter for the Phase-19 thermal-lab comparison.

Renders a thermal-lab comparison (the payload of
``POST /api/thermal-lab/compare`` enriched with a few run-meta keys) into a
multi-sheet ``.xlsx`` workbook:

- **Confronto KPI** — one row per house variant with the headline numbers
  (UA, annual HVAC energy + band, cost, comfort breaches, peak power,
  worst-case indoor temperature, worst days).
- **Serie giornaliera** — the typical-year daily series: outdoor temperature
  plus one HVAC-energy column per variant (365 rows).
- **Temperatura interna** — only when the run used the dynamic RC mode: the
  representative-path daily indoor min/max per variant.

The exporter operates on a plain mapping (no simulation objects) so it stays
trivially unit-testable, mirroring the Phase-11 ``xlsx_cashflow`` exporter.
"""

from __future__ import annotations

from typing import Any, BinaryIO, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_TITLE_FONT = Font(bold=True, size=12)


def _write_header(ws, row: int, columns: list[str]) -> None:
    """Format a row of column headers with the project blue fill."""
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autosize_columns(ws, max_width: int = 30) -> None:
    """Approximate column auto-sizing from the longest string per column."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        widest = 8
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            widest = max(widest, len(str(v)))
        ws.column_dimensions[letter].width = min(max_width, widest + 2)


def build_thermal_lab_xlsx(report: Mapping[str, Any], buffer: BinaryIO) -> None:
    """
    Render a thermal-lab comparison into ``buffer`` as an Excel workbook.

    Args:
        report: A mapping with the keys of the ``/thermal-lab/compare``
            response (``days``, ``daily_outdoor_mean_c``, ``variants``,
            ``n_paths``, ``n_years``) optionally enriched with the run-meta
            keys ``climate_name``, ``dynamic`` and
            ``electricity_price_eur_per_kwh``. Each variant is a mapping with
            the :class:`ThermalVariantResultSchema` fields.
        buffer: Binary writable stream (e.g. an ``io.BytesIO``).

    Raises:
        ValueError: If ``report`` carries no variants — there is nothing to
            export.
    """
    variants = report.get("variants") or []
    if not variants:
        raise ValueError("Thermal-lab report has no variants to export.")

    dynamic = bool(report.get("dynamic"))
    wb = Workbook()

    # ── Sheet 1: KPI comparison ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Confronto KPI"

    ws.cell(row=1, column=1, value="Laboratorio termico — confronto isolamenti").font = _TITLE_FONT
    meta_bits = []
    if report.get("climate_name"):
        meta_bits.append(f"Clima: {report['climate_name']}")
    meta_bits.append(f"{report.get('n_paths', '?')} path × {report.get('n_years', '?')} anno/i")
    meta_bits.append(
        f"Prezzo: {report.get('price_label') or str(report.get('electricity_price_eur_per_kwh', '?')) + ' €/kWh'}"
    )
    meta_bits.append(f"Modello: {'dinamico RC' if dynamic else 'steady-state'}")
    ws.cell(row=2, column=1, value=" · ".join(meta_bits))

    columns = [
        "Configurazione", "UA (kW/°C)", "kWh/anno medio", "kWh/anno p05",
        "kWh/anno p95", "Risc. (kWh/anno)", "Raffr. (kWh/anno)",
        "Costo medio (€/anno)", "Costo p05 (€)", "Costo p95 (€)",
        "Comfort breach (h/anno)", "Picco (kW)", "T int. min (°C)",
        "T int. max (°C)", "Giorno peggiore risc.", "Giorno peggiore raffr.",
    ]
    header_row = 4
    _write_header(ws, row=header_row, columns=columns)

    def _row_values(v: Mapping[str, Any]) -> list[Any]:
        return [
            v.get("label"),
            round(v.get("ua_kw_per_c", 0.0), 4),
            round(v.get("hvac_kwh_annual_mean", 0.0), 1),
            round(v.get("hvac_kwh_annual_p05", 0.0), 1),
            round(v.get("hvac_kwh_annual_p95", 0.0), 1),
            round(v.get("heating_kwh_annual_mean", 0.0), 1),
            round(v.get("cooling_kwh_annual_mean", 0.0), 1),
            round(v.get("annual_cost_eur_mean", 0.0), 1),
            round(v.get("annual_cost_eur_p05", 0.0), 1),
            round(v.get("annual_cost_eur_p95", 0.0), 1),
            round(v.get("comfort_breach_hours_per_year_mean", 0.0), 1),
            round(v.get("p_elec_hvac_peak_kw_mean", 0.0), 2),
            round(v["t_in_min_c"], 1) if dynamic else "—",
            round(v["t_in_max_c"], 1) if dynamic else "—",
            v.get("worst_heating_day_index"),
            v.get("worst_cooling_day_index"),
        ]

    for offset, v in enumerate(variants, start=header_row + 1):
        for col_idx, value in enumerate(_row_values(v), start=1):
            ws.cell(row=offset, column=col_idx, value=value)
    _autosize_columns(ws)

    # ── Sheet 2: typical-year daily series ─────────────────────────────────
    ws_daily = wb.create_sheet("Serie giornaliera")
    labels = [v.get("label", f"Var {i}") for i, v in enumerate(variants)]
    daily_cols = ["Giorno (0-based)", "T esterna (°C)"] + [f"{lbl} (kWh/g)" for lbl in labels]
    _write_header(ws_daily, row=1, columns=daily_cols)

    days = report.get("days") or list(range(len(report.get("daily_outdoor_mean_c") or [])))
    outdoor = report.get("daily_outdoor_mean_c") or []
    series = [v.get("daily_hvac_kwh") or [] for v in variants]
    for r, day in enumerate(days, start=2):
        ws_daily.cell(row=r, column=1, value=day)
        ws_daily.cell(
            row=r, column=2,
            value=round(outdoor[r - 2], 2) if r - 2 < len(outdoor) else None,
        )
        for c, s in enumerate(series, start=3):
            ws_daily.cell(
                row=r, column=c,
                value=round(s[r - 2], 3) if r - 2 < len(s) else None,
            )
    ws_daily.freeze_panes = "A2"
    _autosize_columns(ws_daily)

    # ── Sheet 3: indoor temperature (dynamic mode only) ────────────────────
    has_indoor = dynamic and any(v.get("daily_indoor_min_c") for v in variants)
    if has_indoor:
        ws_t = wb.create_sheet("Temperatura interna")
        cols = ["Giorno (0-based)"]
        for lbl in labels:
            cols += [f"{lbl} T min (°C)", f"{lbl} T max (°C)"]
        _write_header(ws_t, row=1, columns=cols)
        mins = [v.get("daily_indoor_min_c") or [] for v in variants]
        maxs = [v.get("daily_indoor_max_c") or [] for v in variants]
        for r, day in enumerate(days, start=2):
            ws_t.cell(row=r, column=1, value=day)
            col = 2
            for vmin, vmax in zip(mins, maxs):
                ws_t.cell(row=r, column=col,
                          value=round(vmin[r - 2], 2) if r - 2 < len(vmin) else None)
                ws_t.cell(row=r, column=col + 1,
                          value=round(vmax[r - 2], 2) if r - 2 < len(vmax) else None)
                col += 2
        ws_t.freeze_panes = "A2"
        _autosize_columns(ws_t)

    wb.save(buffer)
