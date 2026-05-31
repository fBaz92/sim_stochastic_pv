"""
Excel exporter for the electricity-market lab.

Renders a market-lab run (the payload of ``POST /api/market/run`` enriched with
the scenario run-meta keys) into a multi-sheet ``.xlsx`` workbook:

- **Prezzo medio (mese×ora)** — the display-year wholesale price heatmap, one
  row per calendar month and one column per hour of day (EUR/kWh).
- **Fan annuale** — mean and p05/p95 wholesale price per horizon year.
- **Curva di durata** — the display-year price duration curve.
- **Mix capacita (GW)** — installed capacity per technology over the horizon.
- **Chi fissa il prezzo** — share of the year each technology set the price.

The exporter operates on a plain mapping (no simulation objects) so it stays
trivially unit-testable, mirroring the thermal-lab exporter.
"""

from __future__ import annotations

from typing import Any, BinaryIO, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_TITLE_FONT = Font(bold=True, size=12)

_MONTHS_SHORT = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu",
                 "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]


def _write_header(ws, row: int, columns: list[str]) -> None:
    """Format a row of column headers with the project blue fill."""
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autosize_columns(ws, max_width: int = 30) -> None:
    """Approximate column auto-sizing from the longest string per column."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        widest = 8
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            widest = max(widest, len(str(v)))
        ws.column_dimensions[letter].width = min(max_width, widest + 2)


def build_market_xlsx(report: Mapping[str, Any], buffer: BinaryIO) -> None:
    """
    Render a market-lab run into ``buffer`` as an Excel workbook.

    Args:
        report: A mapping with the keys of the ``/market/run`` response
            (``techs``, ``years``, ``capacity_by_year_gw``, ``display_year``,
            ``price_heatmap_eur_per_kwh``, ``annual_price_*``,
            ``duration_curve_*``, ``price_setter_*``, ``mean_price_eur_per_kwh``,
            ``n_trajectories``, ``n_runs``), optionally enriched with
            ``gas_scenario`` / ``co2_scenario`` / ``coal_scenario``.
        buffer: Binary writable stream (e.g. an ``io.BytesIO``).

    Raises:
        ValueError: If ``report`` carries no price heatmap — nothing to export.
    """
    heatmap = report.get("price_heatmap_eur_per_kwh") or []
    if not heatmap:
        raise ValueError("Market report has no price heatmap to export.")

    display_year = int(report.get("display_year", 0))
    wb = Workbook()

    # ── Sheet 1: price heatmap (month × hour) ──────────────────────────────
    ws = wb.active
    ws.title = "Prezzo medio (mese×ora)"
    ws.cell(row=1, column=1, value="Prezzo all'ingrosso (€/kWh) — mese × ora").font = (
        _TITLE_FONT
    )
    meta_bits = [f"Anno orizzonte: {display_year}"]
    if report.get("gas_scenario"):
        meta_bits.append(f"Gas: {report['gas_scenario']}")
    if report.get("co2_scenario"):
        meta_bits.append(f"CO₂: {report['co2_scenario']}")
    meta_bits.append(f"Prezzo medio: {report.get('mean_price_eur_per_kwh', 0.0):.4f} €/kWh")
    ws.cell(row=2, column=1, value=" · ".join(meta_bits))
    _write_header(ws, 4, ["Mese"] + [f"h{h:02d}" for h in range(24)])
    for m, row in enumerate(heatmap):
        ws.cell(row=5 + m, column=1, value=_MONTHS_SHORT[m] if m < 12 else str(m))
        for h, value in enumerate(row):
            ws.cell(row=5 + m, column=2 + h, value=round(float(value), 5))
    _autosize_columns(ws, max_width=10)

    # ── Sheet 2: annual fan chart data ─────────────────────────────────────
    ws2 = wb.create_sheet("Fan annuale")
    _write_header(ws2, 1, ["Anno", "Media €/kWh", "p05 €/kWh", "p95 €/kWh"])
    years = report.get("years") or []
    mean = report.get("annual_price_mean_eur_per_kwh") or []
    p05 = report.get("annual_price_p05_eur_per_kwh") or []
    p95 = report.get("annual_price_p95_eur_per_kwh") or []
    for i, year in enumerate(years):
        ws2.cell(row=2 + i, column=1, value=int(year))
        ws2.cell(row=2 + i, column=2, value=round(float(mean[i]), 5))
        ws2.cell(row=2 + i, column=3, value=round(float(p05[i]), 5))
        ws2.cell(row=2 + i, column=4, value=round(float(p95[i]), 5))
    _autosize_columns(ws2)

    # ── Sheet 3: duration curve ────────────────────────────────────────────
    ws3 = wb.create_sheet("Curva di durata")
    _write_header(ws3, 1, ["Frazione dell'anno", "Prezzo €/kWh"])
    dur_x = report.get("duration_curve_x") or []
    dur_p = report.get("duration_curve_price_eur_per_kwh") or []
    for i in range(len(dur_x)):
        ws3.cell(row=2 + i, column=1, value=round(float(dur_x[i]), 5))
        ws3.cell(row=2 + i, column=2, value=round(float(dur_p[i]), 5))
    _autosize_columns(ws3)

    # ── Sheet 4: capacity mix over the horizon ─────────────────────────────
    ws4 = wb.create_sheet("Mix capacita (GW)")
    techs = report.get("techs") or []
    cap = report.get("capacity_by_year_gw") or {}
    _write_header(ws4, 1, ["Anno"] + list(techs))
    for i, year in enumerate(years):
        ws4.cell(row=2 + i, column=1, value=int(year))
        for j, tech in enumerate(techs):
            series = cap.get(tech) or []
            if i < len(series):
                ws4.cell(row=2 + i, column=2 + j, value=round(float(series[i]), 3))
    _autosize_columns(ws4)

    # ── Sheet 5: who sets the price (yearly share) ─────────────────────────
    ws5 = wb.create_sheet("Chi fissa il prezzo")
    _write_header(ws5, 1, ["Tecnologia", "Quota dell'anno %"])
    share = report.get("price_setter_share_year") or {}
    for i, (tech, value) in enumerate(
        sorted(share.items(), key=lambda kv: kv[1], reverse=True)
    ):
        ws5.cell(row=2 + i, column=1, value=tech)
        ws5.cell(row=2 + i, column=2, value=round(100.0 * float(value), 2))
    _autosize_columns(ws5)

    wb.save(buffer)
