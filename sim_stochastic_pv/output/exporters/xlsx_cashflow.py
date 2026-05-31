"""
Excel cash-flow exporter (Phase 11).

Renders the mean monthly cash-flow table and the run KPIs into an .xlsx
workbook. Lives in ``output/exporters/`` to keep the API route thin and
to give the CLI a future hook for offline export.

The exporter operates on the ``RunResultRecord.summary`` JSON only — it
never reaches into the on-disk ``output_dir`` (which may have been
deleted) and never re-runs the Monte Carlo simulation.
"""

from __future__ import annotations

from typing import Any, BinaryIO, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _write_header(ws, row: int, columns: list[str]) -> None:
    """Format a row of column headers with the project blue."""
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autosize_columns(ws, max_width: int = 28) -> None:
    """Approximate column auto-sizing based on the longest string per column."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        widest = 8
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            widest = max(widest, len(str(v)))
        ws.column_dimensions[letter].width = min(max_width, widest + 2)


def build_cashflow_xlsx(summary: Mapping[str, Any], buffer: BinaryIO) -> None:
    """
    Render a 2-sheet Excel workbook into ``buffer``.

    Sheet 1 ("Cash flow medio"): one row per month with the mean nominal
    savings, real savings, bonus, cumulative profit (nominal + real),
    electricity price and inflation factor.

    Sheet 2 ("KPI"): the scalar decision metrics — probability of
    profit, IRR, NPV mediano, break-even months, total tax bonus.

    Args:
        summary: The ``RunResultRecord.summary`` JSON. Must contain at
            least ``plots_data.cashflow_table``; the function raises a
            ValueError when it's missing (legacy runs predating Phase 11).
        buffer: Binary writable stream. Typical use: an ``io.BytesIO``
            returned by a FastAPI ``StreamingResponse``.

    Raises:
        ValueError: If the summary does not contain the cash-flow table
            (i.e. the run was generated before Phase 11 wired the payload).
    """
    plots_data = summary.get("plots_data") or {}
    cashflow = plots_data.get("cashflow_table")
    if cashflow is None:
        raise ValueError(
            "Run summary does not include 'plots_data.cashflow_table' — "
            "this run was generated before Phase 11. Re-run the analysis "
            "to get an exportable cash-flow."
        )

    wb = Workbook()

    # ── Sheet 1: monthly cash flow ─────────────────────────────────────
    ws_cf = wb.active
    ws_cf.title = "Cash flow medio"

    columns = [
        ("Mese (0-based)", "months"),
        ("Anno", None),
        ("Mese dell'anno", None),
        ("Risparmio nominale (€)", "mean_savings_eur"),
        ("Risparmio reale (€)", "mean_savings_real_eur"),
        ("Bonus fiscale (€)", "bonus_per_month_eur"),
        ("Immissione (€)", "export_eur"),
        ("Profitto cum. nominale (€)", "mean_profit_cum_eur"),
        ("Profitto cum. reale (€)", "mean_profit_cum_real_eur"),
        ("Prezzo medio (€/kWh)", "mean_price_eur_per_kwh"),
        ("Fattore di inflazione", "mean_inflation_factor"),
    ]
    headers = [c[0] for c in columns]
    _write_header(ws_cf, row=1, columns=headers)

    months = cashflow.get("months", [])
    for row_offset, month in enumerate(months, start=2):
        ws_cf.cell(row=row_offset, column=1, value=month)
        ws_cf.cell(row=row_offset, column=2, value=month // 12)
        ws_cf.cell(row=row_offset, column=3, value=month % 12)
        for col_idx, (_label, key) in enumerate(columns[3:], start=4):
            series = cashflow.get(key, [])
            value = series[row_offset - 2] if row_offset - 2 < len(series) else None
            if isinstance(value, float):
                value = round(value, 4)
            ws_cf.cell(row=row_offset, column=col_idx, value=value)

    ws_cf.freeze_panes = "A2"
    _autosize_columns(ws_cf)

    # ── Sheet 2: KPIs ──────────────────────────────────────────────────
    ws_kpi = wb.create_sheet("KPI")
    _write_header(ws_kpi, row=1, columns=["Metrica", "Valore"])

    def _fmt_pct(p: float | None) -> str | None:
        return None if p is None else f"{p * 100:.1f} %"

    def _fmt_eur(v: float | None) -> str | None:
        return None if v is None else f"€ {v:,.0f}"

    kpi_rows: list[tuple[str, Any]] = [
        ("Scenario", summary.get("scenario")),
        ("Probabilità di guadagno", _fmt_pct(summary.get("prob_gain"))),
        (
            "Probabilità di break-even entro l'orizzonte",
            _fmt_pct(summary.get("prob_break_even_within_horizon")),
        ),
        ("Break-even mediano (mese)", summary.get("break_even_month_median")),
        ("Break-even p05 (mese)", summary.get("break_even_month_p05")),
        ("Break-even p95 (mese)", summary.get("break_even_month_p95")),
        ("IRR atteso (annuo)", _fmt_pct(summary.get("irr_mean"))),
        ("NPV mediano", _fmt_eur(summary.get("npv_median_eur"))),
        ("Guadagno medio nominale (fine orizzonte)", _fmt_eur(summary.get("final_gain_mean_eur"))),
        ("Guadagno medio reale (fine orizzonte)", _fmt_eur(summary.get("final_gain_real_mean_eur"))),
        ("Bonus fiscale totale", _fmt_eur(summary.get("tax_bonus_total_eur"))),
        (
            "Ricavo da immissione totale (medio)",
            _fmt_eur((summary.get("market") or {}).get("export_revenue_total_mean_eur")),
        ),
    ]
    for offset, (label, value) in enumerate(kpi_rows, start=2):
        ws_kpi.cell(row=offset, column=1, value=label)
        ws_kpi.cell(row=offset, column=2, value=value)
    _autosize_columns(ws_kpi, max_width=42)

    wb.save(buffer)
