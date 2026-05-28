"""
Excel templates and parsers for inline load profiles.

The Scenario wizard lets the user define a custom load profile inline
(without going through the Database page). Three shapes are supported:

- ``monthly_avg``  → 12 rows × 1 column   (mean Watt per month)
- ``monthly_24h``  → 12 rows × 24 columns (mean Watt per (month, hour))
- ``weekly``       → 7 rows × 24 columns  (mean Watt per (weekday, hour))
                     plus a 12-row scalar "monthly scale" column

This module exposes:

- ``build_template_xlsx(kind, buffer)`` — writes a blank, well-labelled
  workbook the user can fill in and re-upload.
- ``parse_load_profile_xlsx(kind, buffer)`` — reads a workbook produced
  from the template (or any compatible layout) and returns plain Python
  lists ready to be embedded in the scenario JSON payload.

Both functions are pure and accept ``io.BytesIO``-style streams, so the
API layer can wire them with FastAPI ``StreamingResponse`` / ``UploadFile``
without temporary files.
"""

from __future__ import annotations

from typing import Any, BinaryIO, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


LoadProfileKind = Literal["monthly_avg", "monthly_24h", "weekly"]


_MONTH_LABELS = [
    "Gen", "Feb", "Mar", "Apr", "Mag", "Giu",
    "Lug", "Ago", "Set", "Ott", "Nov", "Dic",
]
_WEEKDAY_LABELS = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_NOTE_FONT = Font(italic=True, color="6C757D")


def _format_header(cell) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _HEADER_ALIGN


def _autosize_columns(ws, max_width: int = 16) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        widest = 8
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            widest = max(widest, len(str(v)))
        ws.column_dimensions[letter].width = min(max_width, widest + 2)


# ─────────────────────────────────────────────────────────────────────
# Template builders
# ─────────────────────────────────────────────────────────────────────


def _build_monthly_avg_template(ws) -> None:
    """Template for ``monthly_avg``: 12 rows × 1 numeric column."""
    ws.title = "Media mensile (W)"
    ws.cell(row=1, column=1, value="Mese")
    ws.cell(row=1, column=2, value="Carico medio (W)")
    _format_header(ws.cell(row=1, column=1))
    _format_header(ws.cell(row=1, column=2))
    for i, label in enumerate(_MONTH_LABELS):
        ws.cell(row=i + 2, column=1, value=label)
        ws.cell(row=i + 2, column=2, value=300.0)  # sensible default placeholder
    ws.freeze_panes = "B2"

    note = ws.cell(
        row=15, column=1,
        value=(
            "Inserisci il consumo medio in Watt per ogni mese. "
            "Il simulatore lo userà come livello base costante 24h/24h."
        ),
    )
    note.font = _NOTE_FONT
    ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column=4)


def _build_monthly_24h_template(ws) -> None:
    """Template for ``monthly_24h``: 12 rows × 24 numeric columns."""
    ws.title = "24h per mese (W)"
    ws.cell(row=1, column=1, value="Mese")
    _format_header(ws.cell(row=1, column=1))
    for h in range(24):
        cell = ws.cell(row=1, column=h + 2, value=f"{h:02d}h")
        _format_header(cell)
    for i, label in enumerate(_MONTH_LABELS):
        ws.cell(row=i + 2, column=1, value=label)
        for h in range(24):
            ws.cell(row=i + 2, column=h + 2, value=200.0)
    ws.freeze_panes = "B2"

    note = ws.cell(
        row=15, column=1,
        value=(
            "Una riga per mese, una colonna per ora (00h–23h). "
            "I valori sono in Watt. Salva il file e ricaricalo dal wizard."
        ),
    )
    note.font = _NOTE_FONT
    ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column=12)


def _build_weekly_template(wb: Workbook) -> None:
    """
    Template for ``weekly``: a 7 × 24 pattern sheet + a 12 × 1 monthly
    scale sheet. The simulator multiplies the weekly base pattern by the
    monthly scale to produce a year-long curve.
    """
    pattern_ws = wb.active
    pattern_ws.title = "Pattern settimanale (W)"
    pattern_ws.cell(row=1, column=1, value="Giorno")
    _format_header(pattern_ws.cell(row=1, column=1))
    for h in range(24):
        c = pattern_ws.cell(row=1, column=h + 2, value=f"{h:02d}h")
        _format_header(c)
    for i, label in enumerate(_WEEKDAY_LABELS):
        pattern_ws.cell(row=i + 2, column=1, value=label)
        for h in range(24):
            pattern_ws.cell(row=i + 2, column=h + 2, value=100.0)
    pattern_ws.freeze_panes = "B2"

    note = pattern_ws.cell(
        row=10, column=1,
        value=(
            "Pattern settimanale: una riga per giorno della settimana "
            "(Lun→Dom), una colonna per ora (00h–23h). Valori in Watt."
        ),
    )
    note.font = _NOTE_FONT
    pattern_ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=12)

    monthly_ws = wb.create_sheet("Scala mensile (W)")
    monthly_ws.cell(row=1, column=1, value="Mese")
    monthly_ws.cell(row=1, column=2, value="Consumo medio (W)")
    _format_header(monthly_ws.cell(row=1, column=1))
    _format_header(monthly_ws.cell(row=1, column=2))
    for i, label in enumerate(_MONTH_LABELS):
        monthly_ws.cell(row=i + 2, column=1, value=label)
        monthly_ws.cell(row=i + 2, column=2, value=300.0)
    monthly_ws.freeze_panes = "B2"

    note2 = monthly_ws.cell(
        row=15, column=1,
        value=(
            "Consumo medio mensile (Watt). Modula la scala globale del "
            "pattern settimanale per ciascun mese dell'anno."
        ),
    )
    note2.font = _NOTE_FONT
    monthly_ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column=4)


def build_template_xlsx(kind: LoadProfileKind, buffer: BinaryIO) -> None:
    """
    Write a blank, well-labelled Excel template into ``buffer``.

    Args:
        kind: Profile shape requested by the user (``monthly_avg``,
            ``monthly_24h`` or ``weekly``).
        buffer: Binary writable stream. Caller is responsible for
            seeking it back to position 0 if streaming downstream.

    Raises:
        ValueError: If ``kind`` is not one of the supported values.
    """
    wb = Workbook()
    if kind == "monthly_avg":
        _build_monthly_avg_template(wb.active)
        _autosize_columns(wb.active)
    elif kind == "monthly_24h":
        _build_monthly_24h_template(wb.active)
        _autosize_columns(wb.active)
    elif kind == "weekly":
        _build_weekly_template(wb)
        for ws in wb.worksheets:
            _autosize_columns(ws)
    else:
        raise ValueError(
            f"Unknown load profile kind '{kind}'. "
            "Expected one of: 'monthly_avg', 'monthly_24h', 'weekly'."
        )
    wb.save(buffer)


# ─────────────────────────────────────────────────────────────────────
# Parsers (workbook → plain Python data)
# ─────────────────────────────────────────────────────────────────────


def _read_numeric_grid(
    ws, *, n_rows: int, n_cols: int, first_row: int = 2, first_col: int = 2
) -> list[list[float]]:
    """Read ``n_rows × n_cols`` numeric values starting at ``(first_row, first_col)``.

    Empty cells become ``0.0``; non-numeric cells raise a ValueError with
    the offending cell coordinate so the user can fix the workbook.
    """
    out: list[list[float]] = []
    for r in range(n_rows):
        row_vals: list[float] = []
        for c in range(n_cols):
            cell = ws.cell(row=first_row + r, column=first_col + c)
            v = cell.value
            if v is None or v == "":
                row_vals.append(0.0)
                continue
            if isinstance(v, bool):
                raise ValueError(
                    f"Cell {cell.coordinate} contains a boolean; expected a number."
                )
            try:
                row_vals.append(float(v))
            except (TypeError, ValueError):
                raise ValueError(
                    f"Cell {cell.coordinate} value {v!r} is not a number."
                )
        out.append(row_vals)
    return out


def parse_load_profile_xlsx(kind: LoadProfileKind, buffer: BinaryIO) -> dict[str, Any]:
    """
    Parse a user-uploaded Excel into a JSON-compatible payload.

    Args:
        kind: Shape expected. Determines how the sheet is read.
        buffer: Binary readable stream pointing at the workbook bytes.

    Returns:
        Dictionary with one of these shapes (matching the wizard state):

        - ``monthly_avg`` → ``{"monthly_avg_w": [12 floats]}``
        - ``monthly_24h`` → ``{"monthly_24h_w": [[24 floats] × 12]}``
        - ``weekly``      → ``{"monthly_w": [12 floats],
                              "weekly_pattern_w": [[24 floats] × 7]}``

    Raises:
        ValueError: On unknown kind or malformed cells.
    """
    wb = load_workbook(buffer, data_only=True)

    if kind == "monthly_avg":
        ws = wb.active
        grid = _read_numeric_grid(ws, n_rows=12, n_cols=1, first_col=2)
        return {"monthly_avg_w": [row[0] for row in grid]}

    if kind == "monthly_24h":
        ws = wb.active
        grid = _read_numeric_grid(ws, n_rows=12, n_cols=24, first_col=2)
        return {"monthly_24h_w": grid}

    if kind == "weekly":
        # Pattern sheet: first sheet, 7 × 24.
        if "Pattern settimanale (W)" in wb.sheetnames:
            pattern_ws = wb["Pattern settimanale (W)"]
        else:
            pattern_ws = wb.worksheets[0]
        pattern = _read_numeric_grid(pattern_ws, n_rows=7, n_cols=24, first_col=2)

        # Monthly scale sheet: second sheet by default.
        if "Scala mensile (W)" in wb.sheetnames:
            monthly_ws = wb["Scala mensile (W)"]
        elif len(wb.worksheets) >= 2:
            monthly_ws = wb.worksheets[1]
        else:
            raise ValueError(
                "Weekly template must contain a second sheet 'Scala mensile (W)' "
                "with one row per month."
            )
        monthly_grid = _read_numeric_grid(monthly_ws, n_rows=12, n_cols=1, first_col=2)

        return {
            "monthly_w": [row[0] for row in monthly_grid],
            "weekly_pattern_w": pattern,
        }

    raise ValueError(
        f"Unknown load profile kind '{kind}'. "
        "Expected one of: 'monthly_avg', 'monthly_24h', 'weekly'."
    )
