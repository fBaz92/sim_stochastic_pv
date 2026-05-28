"""
Phase 11 — Step 8/9 export endpoint tests.

Verify the new endpoints serve a valid Excel workbook (and, in Step 9,
a PDF) for a freshly executed analysis, and that legacy runs (without
the Phase 11 ``cashflow_table`` payload) are rejected with HTTP 422.
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from sim_stochastic_pv.api import dependencies
from sim_stochastic_pv.api.app import create_app
from sim_stochastic_pv.application import SimulationApplication
from sim_stochastic_pv.persistence import PersistenceService


def _build_client(persistence: PersistenceService) -> TestClient:
    app = create_app()

    def get_app_service() -> SimulationApplication:
        return SimulationApplication(
            save_outputs=False, persistence=persistence, result_builder=None,
        )

    app.dependency_overrides[dependencies.get_application_service] = get_app_service
    app.dependency_overrides[dependencies.get_persistence_service] = lambda: persistence
    return TestClient(app)


def _execute_one_analysis(client: TestClient, scenario: dict, *, n_mc: int = 4) -> int:
    """Submit /api/analysis and return the persisted run_id."""
    resp = client.post(
        "/api/analysis",
        json={"n_mc": n_mc, "seed": 7, "scenario": scenario},
    )
    assert resp.status_code == 200, resp.text
    run_id = resp.json()["run_id"]
    assert run_id is not None
    return run_id


def test_export_cashflow_xlsx_returns_workbook(
    persistence: PersistenceService, simple_scenario_data: dict
):
    """A Phase-11 run must export a non-empty 2-sheet Excel workbook."""
    client = _build_client(persistence)
    run_id = _execute_one_analysis(client, simple_scenario_data)

    resp = client.get(f"/api/runs/{run_id}/export/cashflow.xlsx")
    assert resp.status_code == 200, resp.text
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert f'cashflow_run_{run_id}.xlsx' in resp.headers["content-disposition"]
    # First 4 bytes of a valid .xlsx (a zip container) are 'PK\x03\x04'.
    assert resp.content[:4] == b"PK\x03\x04"

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Cash flow medio", "KPI"]
    ws = wb["Cash flow medio"]
    assert ws.cell(row=1, column=1).value == "Mese (0-based)"
    # Data rows beyond the header. The test scenario uses ~few years so
    # we expect at least 12 months populated.
    assert ws.max_row >= 13


def test_export_cashflow_xlsx_404_for_unknown_run(persistence: PersistenceService):
    client = _build_client(persistence)
    resp = client.get("/api/runs/9999/export/cashflow.xlsx")
    assert resp.status_code == 404


def test_export_cashflow_xlsx_422_on_legacy_run(
    persistence: PersistenceService, simple_scenario_data: dict
):
    """A pre-Phase-11 ``summary`` (no cashflow_table) responds with 422."""
    client = _build_client(persistence)
    # Manually persist a legacy-style run record bypassing the analysis
    # flow, then hit the export endpoint.
    legacy_summary = {
        "scenario": "legacy",
        "final_gain_mean_eur": 100.0,
        "final_gain_real_mean_eur": 90.0,
        "prob_gain": 0.5,
        "plots_data": {},  # cashflow_table intentionally absent
    }
    record = persistence.record_run_result("analysis", legacy_summary)
    resp = client.get(f"/api/runs/{record.id}/export/cashflow.xlsx")
    assert resp.status_code == 422
    assert "cashflow_table" in resp.json()["detail"]


def test_export_report_pdf_returns_pdf(
    persistence: PersistenceService, simple_scenario_data: dict
):
    """The PDF endpoint must return a real PDF starting with %PDF-."""
    client = _build_client(persistence)
    run_id = _execute_one_analysis(client, simple_scenario_data)

    resp = client.get(f"/api/runs/{run_id}/export/report.pdf")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert f"report_run_{run_id}.pdf" in resp.headers["content-disposition"]
    assert resp.content[:5] == b"%PDF-"
    # Sanity check: a real report has meaningful size (the cover page +
    # at least one embedded chart should weight well over 5 KB).
    assert len(resp.content) > 5_000


def test_export_report_pdf_404_for_unknown_run(persistence: PersistenceService):
    client = _build_client(persistence)
    resp = client.get("/api/runs/9999/export/report.pdf")
    assert resp.status_code == 404


def test_export_report_pdf_works_on_legacy_run(
    persistence: PersistenceService, simple_scenario_data: dict
):
    """Older summaries (no cashflow_table) still render a smaller PDF."""
    client = _build_client(persistence)
    legacy_summary = {
        "scenario": "legacy",
        "final_gain_mean_eur": 100.0,
        "final_gain_real_mean_eur": 90.0,
        "prob_gain": 0.5,
        "plots_data": {},
    }
    record = persistence.record_run_result("analysis", legacy_summary)
    resp = client.get(f"/api/runs/{record.id}/export/report.pdf")
    assert resp.status_code == 200
    assert resp.content[:5] == b"%PDF-"
