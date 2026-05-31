"""
Tests for the electricity-market lab: orchestrator, API routes, exporters.

The orchestrator tests pin the aggregate shapes/invariants (heatmap, fan bands,
duration-curve monotonicity, capacity trends). The route tests exercise the
thin HTTP layer (run, save/list/delete profiles, exports). The exporter tests
check that the XLSX/PDF builders produce valid bytes from a plain report dict.
"""

from __future__ import annotations

import io

import numpy as np
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from sim_stochastic_pv.api import dependencies
from sim_stochastic_pv.api.app import create_app
from sim_stochastic_pv.output.exporters import build_market_pdf, build_market_xlsx
from sim_stochastic_pv.persistence import PersistenceService
from sim_stochastic_pv.simulation.market_lab import (
    MarketLabConfig,
    TechTrendSpec,
    build_market_provider,
    run_market_lab,
)


def _small_config(**overrides) -> MarketLabConfig:
    base = dict(n_years=3, n_trajectories=3, n_runs=3, seed=0)
    base.update(overrides)
    return MarketLabConfig(**base)


def _small_request_body(**overrides) -> dict:
    body = {"n_years": 3, "n_trajectories": 3, "n_runs": 3, "seed": 0}
    body.update(overrides)
    return body


def _create_test_client(persistence: PersistenceService) -> TestClient:
    app = create_app()
    app.dependency_overrides[dependencies.get_persistence_service] = lambda: persistence
    return TestClient(app)


# ── orchestrator ─────────────────────────────────────────────────────────


def test_run_market_lab_shapes():
    r = run_market_lab(_small_config())
    assert r.price_heatmap_eur_per_kwh.shape == (12, 24)
    assert r.annual_price_mean_eur_per_kwh.shape == (3,)
    assert r.price_setter_dominant.shape == (12, 24)
    assert len(r.techs) == 6
    assert r.mean_price_eur_per_kwh > 0.0


def test_run_market_lab_fan_band_ordered():
    r = run_market_lab(_small_config(n_trajectories=6))
    assert np.all(r.annual_price_p05_eur_per_kwh <= r.annual_price_mean_eur_per_kwh + 1e-9)
    assert np.all(r.annual_price_mean_eur_per_kwh <= r.annual_price_p95_eur_per_kwh + 1e-9)


def test_run_market_lab_duration_curve_descending():
    r = run_market_lab(_small_config())
    p = r.duration_curve_price_eur_per_kwh
    assert np.all(p[:-1] >= p[1:] - 1e-12)
    assert 0.0 <= r.duration_curve_x[0] <= r.duration_curve_x[-1] <= 1.0


def test_run_market_lab_capacity_trend_applied():
    r = run_market_lab(
        _small_config(
            capacities_gw={"solar": 30.0},
            capacity_trends={"solar": TechTrendSpec(annual_growth_pct=10.0)},
        )
    )
    solar = r.capacity_by_year_gw["solar"]
    assert solar[0] == pytest.approx(30.0)
    assert solar[1] == pytest.approx(33.0)  # +10%
    assert solar[2] == pytest.approx(36.3)


def test_run_market_lab_capacity_step_year():
    r = run_market_lab(
        _small_config(
            capacities_gw={"nuclear": 0.0},
            capacity_trends={"nuclear": TechTrendSpec(step_year=2, step_capacity_gw=4.0)},
        )
    )
    nuclear = r.capacity_by_year_gw["nuclear"]
    assert nuclear[0] == pytest.approx(0.0)
    assert nuclear[1] == pytest.approx(0.0)
    assert nuclear[2] == pytest.approx(4.0)


def test_run_market_lab_reproducible():
    a = run_market_lab(_small_config(seed=7))
    b = run_market_lab(_small_config(seed=7))
    assert np.allclose(a.price_heatmap_eur_per_kwh, b.price_heatmap_eur_per_kwh)


def test_price_setter_dominant_indices_valid():
    r = run_market_lab(_small_config())
    dom = r.price_setter_dominant
    n = len(r.price_setter_techs)
    # Every cell is either -1 (none) or a valid index into price_setter_techs.
    assert np.all((dom == -1) | ((dom >= 0) & (dom < n)))


def test_build_market_provider_surface_size():
    provider = build_market_provider(_small_config(), pmg_base_eur_per_kwh=0.05)
    assert provider.price_surface.price_eur_per_kwh.shape == (3, 3, 12, 24)
    assert provider.pmg_base_eur_per_kwh == pytest.approx(0.05)


# ── routes ──────────────────────────────────────────────────────────────


def test_run_endpoint_returns_schema(persistence):
    client = _create_test_client(persistence)
    resp = client.post("/api/market/run", json=_small_request_body())
    assert resp.status_code == 200
    data = resp.json()
    assert len(data["price_heatmap_eur_per_kwh"]) == 12
    assert len(data["price_heatmap_eur_per_kwh"][0]) == 24
    assert len(data["annual_price_mean_eur_per_kwh"]) == 3
    assert set(data["techs"]) >= {"gas", "solar", "wind"}


def test_run_endpoint_rejects_unknown_gas_scenario(persistence):
    client = _create_test_client(persistence)
    resp = client.post("/api/market/run", json=_small_request_body(gas_scenario="bogus"))
    assert resp.status_code == 422


def test_run_endpoint_rejects_unknown_tech(persistence):
    client = _create_test_client(persistence)
    resp = client.post(
        "/api/market/run",
        json=_small_request_body(capacities_gw={"unobtanium": 5.0}),
    )
    assert resp.status_code == 422


def test_profiles_save_list_delete_roundtrip(persistence):
    client = _create_test_client(persistence)
    save = client.post(
        "/api/market/profiles",
        json={
            "name": "Lab Italia",
            "description": "from test",
            "config": _small_request_body(),
            "pmg_base_eur_per_kwh": 0.05,
        },
    )
    assert save.status_code == 200
    profile_id = save.json()["id"]
    assert save.json()["name"] == "Lab Italia"

    listing = client.get("/api/market/profiles")
    assert listing.status_code == 200
    names = [p["name"] for p in listing.json()]
    assert "Lab Italia" in names

    # The saved profile hydrates into a usable provider.
    provider = persistence.load_market_provider(profile_id)
    assert provider is not None
    assert provider.pmg_base_eur_per_kwh == pytest.approx(0.05)

    deleted = client.delete(f"/api/market/profiles/{profile_id}")
    assert deleted.status_code == 200
    assert client.delete(f"/api/market/profiles/{profile_id}").status_code == 404


def test_export_xlsx_endpoint(persistence):
    client = _create_test_client(persistence)
    resp = client.post("/api/market/run/export.xlsx", json=_small_request_body())
    assert resp.status_code == 200
    assert "spreadsheet" in resp.headers["content-type"]
    assert resp.content[:2] == b"PK"  # zip/xlsx magic


def test_export_pdf_endpoint(persistence):
    client = _create_test_client(persistence)
    resp = client.post("/api/market/run/export.pdf", json=_small_request_body())
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:4] == b"%PDF"


# ── exporters ─────────────────────────────────────────────────────────────


def _report() -> dict:
    r = run_market_lab(_small_config())
    return {
        "techs": r.techs,
        "years": r.years,
        "capacity_by_year_gw": {k: v.tolist() for k, v in r.capacity_by_year_gw.items()},
        "display_year": r.display_year,
        "price_heatmap_eur_per_kwh": r.price_heatmap_eur_per_kwh.tolist(),
        "annual_price_mean_eur_per_kwh": r.annual_price_mean_eur_per_kwh.tolist(),
        "annual_price_p05_eur_per_kwh": r.annual_price_p05_eur_per_kwh.tolist(),
        "annual_price_p95_eur_per_kwh": r.annual_price_p95_eur_per_kwh.tolist(),
        "duration_curve_x": r.duration_curve_x.tolist(),
        "duration_curve_price_eur_per_kwh": r.duration_curve_price_eur_per_kwh.tolist(),
        "price_setter_techs": r.price_setter_techs,
        "price_setter_dominant": r.price_setter_dominant.tolist(),
        "price_setter_share_year": r.price_setter_share_year,
        "mean_price_eur_per_kwh": r.mean_price_eur_per_kwh,
        "n_trajectories": r.n_trajectories,
        "n_runs": r.n_runs,
        "gas_scenario": "base",
        "co2_scenario": None,
        "coal_scenario": None,
    }


def test_xlsx_exporter_produces_valid_workbook():
    buf = io.BytesIO()
    build_market_xlsx(_report(), buf)
    buf.seek(0)
    wb = load_workbook(buf)
    assert "Prezzo medio (mese×ora)" in wb.sheetnames
    assert "Fan annuale" in wb.sheetnames


def test_xlsx_exporter_rejects_empty_report():
    with pytest.raises(ValueError):
        build_market_xlsx({"price_heatmap_eur_per_kwh": []}, io.BytesIO())


def test_pdf_exporter_produces_pdf_bytes():
    buf = io.BytesIO()
    build_market_pdf(_report(), buf)
    assert buf.getvalue()[:4] == b"%PDF"
