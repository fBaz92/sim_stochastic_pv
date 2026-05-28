"""
Tests for the Excel template / parse round-trip of inline load profiles.

Covers the three supported shapes (``monthly_avg``, ``monthly_24h``,
``weekly``), the FastAPI endpoints that wrap them, and the error
handling on malformed inputs.
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from sim_stochastic_pv.api import dependencies
from sim_stochastic_pv.api.app import create_app
from sim_stochastic_pv.application import SimulationApplication
from sim_stochastic_pv.output.exporters import (
    build_template_xlsx,
    parse_load_profile_xlsx,
)
from sim_stochastic_pv.persistence import PersistenceService


def _client(persistence: PersistenceService) -> TestClient:
    app = create_app()
    app.dependency_overrides[dependencies.get_application_service] = (
        lambda: SimulationApplication(save_outputs=False, persistence=persistence)
    )
    app.dependency_overrides[dependencies.get_persistence_service] = lambda: persistence
    return TestClient(app)


# ─────────────────────────────────────────────────────────────
# Direct module tests (no FastAPI in the loop)
# ─────────────────────────────────────────────────────────────


class TestTemplateAndParseRoundTrip:
    def test_monthly_avg_round_trip(self):
        buf = io.BytesIO()
        build_template_xlsx("monthly_avg", buf)
        buf.seek(0)
        # Overwrite the default placeholders with specific values.
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        for i in range(12):
            ws.cell(row=2 + i, column=2, value=100.0 + i)
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        parsed = parse_load_profile_xlsx("monthly_avg", out)
        assert parsed == {"monthly_avg_w": [100.0 + i for i in range(12)]}

    def test_monthly_24h_round_trip(self):
        buf = io.BytesIO()
        build_template_xlsx("monthly_24h", buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        for m in range(12):
            for h in range(24):
                ws.cell(row=2 + m, column=2 + h, value=float(m * 100 + h))
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        parsed = parse_load_profile_xlsx("monthly_24h", out)
        assert parsed.keys() == {"monthly_24h_w"}
        grid = parsed["monthly_24h_w"]
        assert len(grid) == 12
        assert all(len(row) == 24 for row in grid)
        assert grid[5][3] == pytest.approx(503.0)

    def test_weekly_round_trip(self):
        buf = io.BytesIO()
        build_template_xlsx("weekly", buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        pattern_ws = wb["Pattern settimanale (W)"]
        monthly_ws = wb["Scala mensile (W)"]
        for d in range(7):
            for h in range(24):
                pattern_ws.cell(row=2 + d, column=2 + h, value=float(d * 100 + h))
        for m in range(12):
            monthly_ws.cell(row=2 + m, column=2, value=float(200 + m))
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        parsed = parse_load_profile_xlsx("weekly", out)
        assert parsed.keys() == {"monthly_w", "weekly_pattern_w"}
        assert parsed["monthly_w"] == [float(200 + m) for m in range(12)]
        assert len(parsed["weekly_pattern_w"]) == 7
        assert parsed["weekly_pattern_w"][3][5] == pytest.approx(305.0)


class TestParseErrorHandling:
    def test_unknown_kind_raises(self):
        buf = io.BytesIO()
        build_template_xlsx("monthly_avg", buf)
        buf.seek(0)
        with pytest.raises(ValueError, match="Unknown load profile kind"):
            parse_load_profile_xlsx("nonsense", buf)  # type: ignore[arg-type]

    def test_non_numeric_cell_raises_with_coordinate(self):
        buf = io.BytesIO()
        build_template_xlsx("monthly_avg", buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        wb.active.cell(row=4, column=2, value="not a number")  # row 4 = March
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        with pytest.raises(ValueError, match="B4"):
            parse_load_profile_xlsx("monthly_avg", out)


# ─────────────────────────────────────────────────────────────
# Endpoint integration tests
# ─────────────────────────────────────────────────────────────


class TestEndpoints:
    def test_template_download(self, persistence: PersistenceService):
        client = _client(persistence)
        resp = client.get("/api/load-profiles/template/monthly_24h.xlsx")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats"
        )
        assert "load_profile_monthly_24h_template.xlsx" in resp.headers["content-disposition"]
        assert resp.content[:4] == b"PK\x03\x04"  # zip magic

    def test_template_unknown_kind_404(self, persistence: PersistenceService):
        client = _client(persistence)
        resp = client.get("/api/load-profiles/template/something_else.xlsx")
        assert resp.status_code == 404

    def test_upload_round_trip(self, persistence: PersistenceService):
        client = _client(persistence)
        # Build a workbook in memory and POST it.
        buf = io.BytesIO()
        build_template_xlsx("monthly_24h", buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        for m in range(12):
            for h in range(24):
                ws.cell(row=2 + m, column=2 + h, value=42.0)
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        resp = client.post(
            "/api/load-profiles/parse-xlsx/monthly_24h",
            files={
                "file": (
                    "profile.xlsx",
                    out.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 200, resp.text
        data = resp.json()
        assert data.keys() == {"monthly_24h_w"}
        grid = data["monthly_24h_w"]
        assert len(grid) == 12
        assert all(v == 42.0 for row in grid for v in row)

    def test_upload_malformed_returns_422(self, persistence: PersistenceService):
        client = _client(persistence)
        buf = io.BytesIO()
        build_template_xlsx("monthly_avg", buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        wb.active.cell(row=3, column=2, value="oops")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        resp = client.post(
            "/api/load-profiles/parse-xlsx/monthly_avg",
            files={
                "file": (
                    "bad.xlsx",
                    out.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 422
        assert "B3" in resp.json()["detail"]
