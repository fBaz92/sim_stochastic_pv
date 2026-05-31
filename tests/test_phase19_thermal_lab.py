"""
Phase 19 — thermal-lab tests.

Covers the three deliverables of the backend foundation slice:

1. Time-of-day setpoint schedules in :class:`HvacController` (the seam
   prepared in Phase 18): schedules change energy, validate length and the
   per-hour dead-band invariant, and reduce *exactly* to the scalar when set
   to a constant.
2. The :mod:`sim_stochastic_pv.simulation.thermal_lab` comparison engine:
   monotonic energy in insulation, cost = kWh × price, reproducibility,
   dynamic-mode under-sizing drift, worst-day classification, timeseries.
3. The ``/api/thermal-lab/{compare,timeseries}`` endpoints.

All tests are deterministic (fixed seeds) and keep ``n_paths`` small so the
suite stays fast.
"""

from __future__ import annotations

import numpy as np
import pytest
from fastapi.testclient import TestClient

from sim_stochastic_pv.api import dependencies
from sim_stochastic_pv.api.app import create_app
from sim_stochastic_pv.persistence import PersistenceService
from sim_stochastic_pv.persistence.climate_repo import serialize_thermal_model
from sim_stochastic_pv.simulation.thermal import (
    HarmonicSeasonalMean,
    ThermalModel,
    ThermalMonthParams,
)
from sim_stochastic_pv.simulation.thermal_lab import (
    HouseVariant,
    ThermalLabConfig,
    compare_house_variants,
    simulate_thermal_timeseries,
)
from sim_stochastic_pv.simulation.thermal_load import (
    HeatPumpConfig,
    HouseThermalConfig,
    HvacController,
    SetpointConfig,
    ThermalLoadConfig,
)


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


def _build_model(a0: float = 8.0, a1: float = -11.0) -> ThermalModel:
    """A simple stationary climate. Default is a cold Italian-ish winter
    (annual mean 8 °C, ±11 °C seasonal swing, trough at January 1)."""
    harmonic = HarmonicSeasonalMean(a0=a0, a1=a1, a2=0.0)
    params = [
        ThermalMonthParams(
            t_std_residual_c=2.0,
            persistence_phi=0.7,
            t_amplitude_c=5.0,
        )
        for _ in range(12)
    ]
    return ThermalModel(harmonic, params)


def _save_climate_profile(
    persistence: PersistenceService, model: ThermalModel, name: str = "TestClimate"
) -> int:
    """Serialize a thermal model and upsert it as a climate profile; return id."""
    blob = serialize_thermal_model(model)
    record = persistence.climate.upsert_climate_profile(
        {
            "name": name,
            "location_name": "Test Location",
            "latitude": 44.0,
            "longitude": 11.0,
            "harmonic": blob["harmonic"],
            "monthly_params": blob["monthly_params"],
            "climate_trend_c_per_year": blob["climate_trend_c_per_year"],
        }
    )
    return record.id


def _create_test_client(persistence: PersistenceService) -> TestClient:
    app = create_app()
    app.dependency_overrides[dependencies.get_persistence_service] = lambda: persistence
    return TestClient(app)


def _three_variants() -> tuple[HouseVariant, ...]:
    return (
        HouseVariant("poor", HouseThermalConfig(insulation_preset="poor")),
        HouseVariant("standard", HouseThermalConfig(insulation_preset="standard")),
        HouseVariant("good", HouseThermalConfig(insulation_preset="good")),
    )


# ---------------------------------------------------------------------------
# 1. Time-of-day setpoint schedule
# ---------------------------------------------------------------------------


class TestSetpointSchedule:
    def test_constant_schedule_is_byte_identical_to_scalar(self) -> None:
        """A 24-entry schedule equal to the scalar must produce identical
        electric draw — proving the schedule machinery reduces to the scalar."""
        model = _build_model()
        t_hourly = model.to_hourly(model.simulate_daily_means(60, np.random.default_rng(1)))
        house = HouseThermalConfig(insulation_preset="standard")
        hp = HeatPumpConfig(p_elec_max_kw=5.0)

        scalar = HvacController(
            ThermalLoadConfig(enabled=True, house=house, heat_pump=hp,
                              setpoint=SetpointConfig(t_setpoint_heating_c=20.0,
                                                      t_setpoint_cooling_c=26.0))
        )
        scheduled = HvacController(
            ThermalLoadConfig(enabled=True, house=house, heat_pump=hp,
                              setpoint=SetpointConfig(heating_schedule_c=(20.0,) * 24,
                                                      cooling_schedule_c=(26.0,) * 24))
        )
        p_scalar, _ = scalar.compute_hourly_p_elec_kw(t_hourly)
        p_sched, _ = scheduled.compute_hourly_p_elec_kw(t_hourly)
        np.testing.assert_array_equal(p_scalar, p_sched)

    def test_night_setback_reduces_heating_energy(self) -> None:
        """A schedule that drops the night heating setpoint must use less
        energy than holding 20 °C all day in a cold climate."""
        model = _build_model()
        t_hourly = model.to_hourly(model.simulate_daily_means(365, np.random.default_rng(2)))
        house = HouseThermalConfig(insulation_preset="poor")
        hp = HeatPumpConfig(p_elec_max_kw=10.0)

        constant = HvacController(
            ThermalLoadConfig(enabled=True, house=house, heat_pump=hp,
                              setpoint=SetpointConfig())
        )
        # Heating 17 °C from 23:00–06:00, 20 °C otherwise.
        night = [17.0 if (h < 6 or h >= 23) else 20.0 for h in range(24)]
        setback = HvacController(
            ThermalLoadConfig(enabled=True, house=house, heat_pump=hp,
                              setpoint=SetpointConfig(heating_schedule_c=tuple(night)))
        )
        e_const = constant.compute_hourly_p_elec_kw(t_hourly)[0].sum()
        e_setback = setback.compute_hourly_p_elec_kw(t_hourly)[0].sum()
        assert e_setback < e_const

    def test_schedule_wrong_length_raises(self) -> None:
        with pytest.raises(ValueError, match="24 hour-of-day"):
            SetpointConfig(heating_schedule_c=(20.0,) * 23)

    def test_schedule_deadband_invariant_per_hour(self) -> None:
        """Heating ≥ cooling at any hour must be rejected."""
        bad_heating = [20.0] * 24
        bad_heating[12] = 30.0  # exceeds the 26 °C cooling setpoint at noon
        with pytest.raises(ValueError, match="hour 12"):
            SetpointConfig(heating_schedule_c=tuple(bad_heating))

    def test_setpoint_arrays_indexed_by_hour_of_day(self) -> None:
        """The public accessor must repeat the 24-entry schedule per day."""
        # Distinct value per hour, all safely below the 26 °C cooling default.
        ctrl = HvacController(
            ThermalLoadConfig(
                enabled=True,
                setpoint=SetpointConfig(
                    heating_schedule_c=tuple(12.0 + 0.1 * h for h in range(24))
                ),
            )
        )
        at_home = np.ones(48, dtype=bool)  # two days
        heat, _cool = ctrl.setpoint_arrays(at_home)
        # Hour 5 of day 0 and hour 5 of day 1 share the same scheduled value.
        assert heat[5] == pytest.approx(12.5)
        assert heat[24 + 5] == pytest.approx(12.5)


# ---------------------------------------------------------------------------
# 2. Comparison engine
# ---------------------------------------------------------------------------


class TestCompareHouseVariants:
    def test_energy_monotonic_in_insulation(self) -> None:
        model = _build_model()
        cfg = ThermalLabConfig(
            house_variants=_three_variants(),
            heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
            setpoint=SetpointConfig(),
        )
        res = compare_house_variants(model, cfg, n_paths=12, n_years=1, seed=0)
        poor, standard, good = res.variants
        assert poor.hvac_kwh_annual_mean > standard.hvac_kwh_annual_mean > good.hvac_kwh_annual_mean
        assert poor.ua_kw_per_c > standard.ua_kw_per_c > good.ua_kw_per_c

    def test_cost_equals_kwh_times_price(self) -> None:
        model = _build_model()
        cfg = ThermalLabConfig(
            house_variants=_three_variants()[:1],
            heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
            setpoint=SetpointConfig(),
            electricity_price_eur_per_kwh=0.30,
        )
        res = compare_house_variants(model, cfg, n_paths=10, seed=0)
        v = res.variants[0]
        assert v.annual_cost_eur_mean == pytest.approx(v.hvac_kwh_annual_mean * 0.30)
        assert v.annual_cost_eur_p95 == pytest.approx(v.hvac_kwh_annual_p95 * 0.30)

    def test_reproducible_with_seed(self) -> None:
        model = _build_model()
        cfg = ThermalLabConfig(
            house_variants=_three_variants(),
            heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
            setpoint=SetpointConfig(),
        )
        a = compare_house_variants(model, cfg, n_paths=8, seed=123)
        b = compare_house_variants(model, cfg, n_paths=8, seed=123)
        for va, vb in zip(a.variants, b.variants):
            assert va.hvac_kwh_annual_mean == vb.hvac_kwh_annual_mean
            np.testing.assert_array_equal(va.daily_hvac_kwh, vb.daily_hvac_kwh)

    def test_daily_series_shapes(self) -> None:
        model = _build_model()
        cfg = ThermalLabConfig(
            house_variants=_three_variants()[:1],
            heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
            setpoint=SetpointConfig(),
        )
        res = compare_house_variants(model, cfg, n_paths=5, n_years=2, seed=0)
        assert res.days.shape == (365,)
        assert res.daily_outdoor_mean_c.shape == (365,)
        assert res.variants[0].daily_hvac_kwh.shape == (365,)

    def test_dynamic_undersized_pump_drifts_below_setpoint(self) -> None:
        """A poorly insulated house with a tiny heat pump cannot hold 20 °C:
        dynamic mode must report indoor temp below setpoint + comfort breaches."""
        model = _build_model()
        cfg = ThermalLabConfig(
            house_variants=(HouseVariant("poor", HouseThermalConfig(insulation_preset="poor")),),
            heat_pump=HeatPumpConfig(p_elec_max_kw=1.0),  # undersized
            setpoint=SetpointConfig(t_setpoint_heating_c=20.0),
            dynamic=True,
        )
        res = compare_house_variants(model, cfg, n_paths=6, seed=0)
        v = res.variants[0]
        assert v.t_in_min_c < 20.0
        assert v.comfort_breach_hours_per_year_mean > 0.0
        assert v.daily_indoor_min_c is not None
        assert v.daily_indoor_min_c.shape == (365,)

    def test_steady_state_has_no_indoor_series(self) -> None:
        model = _build_model()
        cfg = ThermalLabConfig(
            house_variants=_three_variants()[:1],
            heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
            setpoint=SetpointConfig(),
            dynamic=False,
        )
        res = compare_house_variants(model, cfg, n_paths=4, seed=0)
        assert res.variants[0].daily_indoor_min_c is None

    def test_worst_days_cold_climate(self) -> None:
        """Cold climate: a worst heating day in winter, no cooling day."""
        model = _build_model(a0=6.0, a1=-12.0)
        cfg = ThermalLabConfig(
            house_variants=_three_variants()[:1],
            heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
            setpoint=SetpointConfig(),
        )
        res = compare_house_variants(model, cfg, n_paths=10, seed=0)
        v = res.variants[0]
        assert v.worst_heating_day_index is not None
        # Winter: first or last ~5 weeks of the calendar year.
        assert v.worst_heating_day_index < 40 or v.worst_heating_day_index > 325
        assert v.worst_cooling_day_index is None  # never warm enough to cool

    def test_worst_days_hot_climate_has_cooling(self) -> None:
        """Hot climate: both a winter heating day and a summer cooling day."""
        model = _build_model(a0=22.0, a1=-12.0)
        cfg = ThermalLabConfig(
            house_variants=_three_variants()[:1],
            heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
            setpoint=SetpointConfig(),
        )
        res = compare_house_variants(model, cfg, n_paths=10, seed=0)
        v = res.variants[0]
        assert v.worst_cooling_day_index is not None
        # Summer peak around day ~182 (July).
        assert 150 < v.worst_cooling_day_index < 215

    def test_empty_variants_raises(self) -> None:
        with pytest.raises(ValueError, match="at least one variant"):
            ThermalLabConfig(
                house_variants=(),
                heat_pump=HeatPumpConfig(),
                setpoint=SetpointConfig(),
            )

    def test_invalid_home_hours_raises(self) -> None:
        with pytest.raises(ValueError, match="range\\(24\\)"):
            ThermalLabConfig(
                house_variants=_three_variants()[:1],
                heat_pump=HeatPumpConfig(),
                setpoint=SetpointConfig(),
                home_hours_of_day=(8, 25),
            )

    def test_n_paths_must_be_positive(self) -> None:
        model = _build_model()
        cfg = ThermalLabConfig(
            house_variants=_three_variants()[:1],
            heat_pump=HeatPumpConfig(),
            setpoint=SetpointConfig(),
        )
        with pytest.raises(ValueError, match="n_paths"):
            compare_house_variants(model, cfg, n_paths=0)


# ---------------------------------------------------------------------------
# 2b. Timeseries
# ---------------------------------------------------------------------------


class TestTimeseries:
    def test_dynamic_returns_indoor_trajectory(self) -> None:
        model = _build_model()
        ts = simulate_thermal_timeseries(
            model,
            house=HouseThermalConfig(insulation_preset="poor"),
            heat_pump=HeatPumpConfig(p_elec_max_kw=3.0),
            setpoint=SetpointConfig(),
            dynamic=True,
            n_days=10,
            seed=3,
        )
        assert ts.hours.shape == (240,)
        assert ts.t_outdoor_c.shape == (240,)
        assert ts.t_indoor_c is not None and ts.t_indoor_c.shape == (240,)
        assert ts.p_elec_hvac_kw.shape == (240,)

    def test_steady_state_has_no_indoor(self) -> None:
        model = _build_model()
        ts = simulate_thermal_timeseries(
            model,
            house=HouseThermalConfig(insulation_preset="standard"),
            heat_pump=HeatPumpConfig(p_elec_max_kw=5.0),
            setpoint=SetpointConfig(),
            dynamic=False,
            n_days=5,
            seed=1,
        )
        assert ts.t_indoor_c is None

    def test_away_hours_yield_infinite_setpoints(self) -> None:
        """With occupancy restricted and no away setback, away hours get
        ±inf setpoints (which the route serialises as null)."""
        model = _build_model()
        ts = simulate_thermal_timeseries(
            model,
            house=HouseThermalConfig(insulation_preset="standard"),
            heat_pump=HeatPumpConfig(p_elec_max_kw=5.0),
            setpoint=SetpointConfig(t_setpoint_away_c=None),
            dynamic=False,
            home_hours_of_day=list(range(7, 22)),  # away overnight
            n_days=2,
            seed=1,
        )
        assert np.isneginf(ts.t_set_heating_c).any()
        assert np.isposinf(ts.t_set_cooling_c).any()


# ---------------------------------------------------------------------------
# 3. API endpoints
# ---------------------------------------------------------------------------


class TestThermalLabAPI:
    def test_compare_endpoint_returns_schema(self, persistence) -> None:
        profile_id = _save_climate_profile(persistence, _build_model())
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/compare",
            json={
                "climate_profile_id": profile_id,
                "n_paths": 8,
                "seed": 0,
                "heat_pump": {"p_elec_max_kw": 8.0},
                "house_variants": [
                    {"label": "poor", "insulation_preset": "poor"},
                    {"label": "good", "insulation_preset": "good"},
                ],
            },
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert len(body["days"]) == 365
        assert len(body["daily_outdoor_mean_c"]) == 365
        assert len(body["variants"]) == 2
        poor, good = body["variants"]
        assert poor["hvac_kwh_annual_mean"] > good["hvac_kwh_annual_mean"]
        assert len(poor["daily_hvac_kwh"]) == 365

    def test_compare_endpoint_404_missing_profile(self, persistence) -> None:
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/compare",
            json={
                "climate_profile_id": 9999,
                "house_variants": [{"label": "x", "insulation_preset": "good"}],
            },
        )
        assert resp.status_code == 404

    def test_compare_endpoint_400_invalid_preset(self, persistence) -> None:
        profile_id = _save_climate_profile(persistence, _build_model())
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/compare",
            json={
                "climate_profile_id": profile_id,
                "house_variants": [{"label": "x", "insulation_preset": "nonsense"}],
            },
        )
        assert resp.status_code == 400
        assert "preset" in resp.json()["detail"].lower()

    def test_timeseries_endpoint_dynamic(self, persistence) -> None:
        profile_id = _save_climate_profile(persistence, _build_model())
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/timeseries",
            json={
                "climate_profile_id": profile_id,
                "n_days": 7,
                "dynamic": True,
                "house": {"insulation_preset": "poor"},
                "heat_pump": {"p_elec_max_kw": 3.0},
            },
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert len(body["hours"]) == 7 * 24
        assert body["t_indoor_c"] is not None
        assert len(body["t_indoor_c"]) == 7 * 24

    def test_timeseries_endpoint_away_setpoints_null(self, persistence) -> None:
        profile_id = _save_climate_profile(persistence, _build_model())
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/timeseries",
            json={
                "climate_profile_id": profile_id,
                "n_days": 2,
                "dynamic": False,
                "home_hours_of_day": list(range(7, 22)),
                "house": {"insulation_preset": "standard"},
            },
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        # Overnight away hours have no setback ⇒ null setpoints in the payload.
        assert any(x is None for x in body["t_set_heating_c"])


# ---------------------------------------------------------------------------
# 4. Report exporters (Excel + PDF)
# ---------------------------------------------------------------------------


def _sample_report() -> dict:
    """A minimal compare-response-shaped mapping for the exporters."""
    return {
        "days": list(range(365)),
        "daily_outdoor_mean_c": [10.0] * 365,
        "n_paths": 8,
        "n_years": 1,
        "climate_name": "Test Location",
        "dynamic": True,
        "electricity_price_eur_per_kwh": 0.25,
        "variants": [
            {
                "label": "poor", "ua_kw_per_c": 0.25,
                "hvac_kwh_annual_mean": 5000.0, "hvac_kwh_annual_p05": 4800.0,
                "hvac_kwh_annual_p95": 5200.0, "annual_cost_eur_mean": 1250.0,
                "annual_cost_eur_p05": 1200.0, "annual_cost_eur_p95": 1300.0,
                "comfort_breach_hours_per_year_mean": 12.0,
                "p_elec_hvac_peak_kw_mean": 2.5, "t_in_min_c": 18.0, "t_in_max_c": 26.0,
                "daily_hvac_kwh": [13.7] * 365,
                "daily_indoor_min_c": [19.0] * 365, "daily_indoor_max_c": [21.0] * 365,
                "worst_heating_day_index": 10, "worst_cooling_day_index": None,
            },
            {
                "label": "good", "ua_kw_per_c": 0.08,
                "hvac_kwh_annual_mean": 1800.0, "hvac_kwh_annual_p05": 1700.0,
                "hvac_kwh_annual_p95": 1900.0, "annual_cost_eur_mean": 450.0,
                "annual_cost_eur_p05": 425.0, "annual_cost_eur_p95": 475.0,
                "comfort_breach_hours_per_year_mean": 0.0,
                "p_elec_hvac_peak_kw_mean": 0.8, "t_in_min_c": 20.0, "t_in_max_c": 26.0,
                "daily_hvac_kwh": [4.9] * 365,
                "daily_indoor_min_c": [20.0] * 365, "daily_indoor_max_c": [22.0] * 365,
                "worst_heating_day_index": 10, "worst_cooling_day_index": None,
            },
        ],
    }


class TestThermalLabExporters:
    def test_xlsx_exporter_produces_valid_workbook(self) -> None:
        import io
        from openpyxl import load_workbook
        from sim_stochastic_pv.output.exporters import build_thermal_lab_xlsx

        buf = io.BytesIO()
        build_thermal_lab_xlsx(_sample_report(), buf)
        buf.seek(0)
        wb = load_workbook(buf)
        assert "Confronto KPI" in wb.sheetnames
        assert "Serie giornaliera" in wb.sheetnames
        # Dynamic report ⇒ indoor-temperature sheet present.
        assert "Temperatura interna" in wb.sheetnames
        # 365 data rows + 1 header on the daily sheet.
        assert wb["Serie giornaliera"].max_row == 366

    def test_xlsx_exporter_rejects_empty_report(self) -> None:
        import io
        from sim_stochastic_pv.output.exporters import build_thermal_lab_xlsx

        with pytest.raises(ValueError, match="no variants"):
            build_thermal_lab_xlsx({"variants": []}, io.BytesIO())

    def test_pdf_exporter_produces_pdf_bytes(self) -> None:
        import io
        from sim_stochastic_pv.output.exporters import build_thermal_lab_pdf

        buf = io.BytesIO()
        build_thermal_lab_pdf(_sample_report(), buf)
        buf.seek(0)
        head = buf.read(5)
        assert head == b"%PDF-", "output is not a PDF"

    def test_compare_export_xlsx_endpoint(self, persistence) -> None:
        profile_id = _save_climate_profile(persistence, _build_model())
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/compare/export.xlsx",
            json={
                "climate_profile_id": profile_id,
                "n_paths": 6,
                "seed": 0,
                "house_variants": [{"label": "poor", "insulation_preset": "poor"}],
            },
        )
        assert resp.status_code == 200, resp.text
        assert "spreadsheetml" in resp.headers["content-type"]
        # XLSX files are ZIP archives → magic bytes "PK".
        assert resp.content[:2] == b"PK"

    def test_compare_export_xlsx_404(self, persistence) -> None:
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/compare/export.xlsx",
            json={
                "climate_profile_id": 9999,
                "house_variants": [{"label": "x", "insulation_preset": "good"}],
            },
        )
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# 5. Phase 19-bis — price coupling, heating/cooling split, preview window
# ---------------------------------------------------------------------------


def _one_variant_config(**kwargs) -> ThermalLabConfig:
    return ThermalLabConfig(
        house_variants=(HouseVariant("poor", HouseThermalConfig(insulation_preset="poor")),),
        heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
        setpoint=SetpointConfig(),
        **kwargs,
    )


class TestPhase19BisPriceCoupling:
    def test_scalar_cost_unchanged_without_price_model(self) -> None:
        """No price model ⇒ cost is exactly annual kWh × scalar price."""
        model = _build_model()
        cfg = _one_variant_config(electricity_price_eur_per_kwh=0.30)
        res = compare_house_variants(model, cfg, n_paths=12, seed=0)
        v = res.variants[0]
        assert v.annual_cost_eur_mean == pytest.approx(v.hvac_kwh_annual_mean * 0.30)

    def test_gbm_widens_cost_band(self) -> None:
        """A volatile GBM price spreads the cost p05–p95 band far beyond the
        (energy-only) scalar band."""
        from sim_stochastic_pv.simulation.prices import GBMPriceModel

        model = _build_model()
        cfg = _one_variant_config(electricity_price_eur_per_kwh=0.25)
        scalar = compare_house_variants(model, cfg, n_paths=40, n_years=5, seed=0)
        gbm_model = GBMPriceModel(
            base_price_eur_per_kwh=0.25, drift_annual=0.03, volatility_annual=0.25
        )
        gbm = compare_house_variants(
            model, cfg, n_paths=40, n_years=5, seed=0, price_model=gbm_model
        )
        sv = scalar.variants[0]
        gv = gbm.variants[0]
        scalar_width = sv.annual_cost_eur_p95 - sv.annual_cost_eur_p05
        gbm_width = gv.annual_cost_eur_p95 - gv.annual_cost_eur_p05
        assert gbm_width > 3.0 * scalar_width

    def test_price_model_reproducible_with_seed(self) -> None:
        from sim_stochastic_pv.simulation.prices import GBMPriceModel

        model = _build_model()
        cfg = _one_variant_config()
        make = lambda: GBMPriceModel(  # noqa: E731
            base_price_eur_per_kwh=0.25, drift_annual=0.03, volatility_annual=0.2
        )
        a = compare_house_variants(model, cfg, n_paths=10, n_years=3, seed=7, price_model=make())
        b = compare_house_variants(model, cfg, n_paths=10, n_years=3, seed=7, price_model=make())
        assert a.variants[0].annual_cost_eur_mean == b.variants[0].annual_cost_eur_mean
        assert a.variants[0].annual_cost_eur_p95 == b.variants[0].annual_cost_eur_p95

    def test_escalating_price_raises_mean_cost(self) -> None:
        """A deterministic escalation over several years lifts the mean annual
        cost above the flat-price case."""
        from sim_stochastic_pv.simulation.prices import EscalatingPriceModel

        model = _build_model()
        cfg = _one_variant_config(electricity_price_eur_per_kwh=0.25)
        flat = compare_house_variants(model, cfg, n_paths=20, n_years=10, seed=0)
        esc_model = EscalatingPriceModel(
            base_price_eur_per_kwh=0.25,
            annual_escalation=0.05,
            use_stochastic_escalation=False,
        )
        esc = compare_house_variants(
            model, cfg, n_paths=20, n_years=10, seed=0, price_model=esc_model
        )
        assert esc.variants[0].annual_cost_eur_mean > flat.variants[0].annual_cost_eur_mean


class TestPhase19BisHeatingCoolingSplit:
    def test_split_sums_to_total_in_steady_state(self) -> None:
        model = _build_model()
        cfg = _one_variant_config()
        res = compare_house_variants(model, cfg, n_paths=10, seed=0)
        v = res.variants[0]
        assert v.heating_kwh_annual_mean + v.cooling_kwh_annual_mean == pytest.approx(
            v.hvac_kwh_annual_mean, rel=1e-6
        )

    def test_cold_climate_is_heating_dominated(self) -> None:
        """A cold climate is overwhelmingly heating — cooling is negligible
        (only rare warm summer afternoons ever exceed the cooling setpoint)."""
        model = _build_model(a0=6.0, a1=-12.0)
        cfg = _one_variant_config()
        res = compare_house_variants(model, cfg, n_paths=10, seed=0)
        v = res.variants[0]
        assert v.heating_kwh_annual_mean > 0.0
        assert v.cooling_kwh_annual_mean < 0.01 * v.heating_kwh_annual_mean

    def test_hot_climate_has_cooling(self) -> None:
        model = _build_model(a0=24.0, a1=-10.0)
        cfg = _one_variant_config()
        res = compare_house_variants(model, cfg, n_paths=10, seed=0)
        v = res.variants[0]
        assert v.cooling_kwh_annual_mean > 0.0


class TestPhase19BisPreviewWindow:
    def test_summer_window_warmer_than_winter(self) -> None:
        model = _build_model(a0=22.0, a1=-12.0)
        common = dict(
            house=HouseThermalConfig(insulation_preset="poor"),
            heat_pump=HeatPumpConfig(p_elec_max_kw=8.0),
            setpoint=SetpointConfig(),
            dynamic=False,
            n_days=10,
            seed=1,
        )
        winter = simulate_thermal_timeseries(model, start_day=0, **common)
        summer = simulate_thermal_timeseries(model, start_day=181, **common)
        assert summer.t_outdoor_c.mean() > winter.t_outdoor_c.mean() + 5.0

    def test_invalid_start_day_raises(self) -> None:
        model = _build_model()
        with pytest.raises(ValueError, match="start_day"):
            simulate_thermal_timeseries(
                model,
                house=HouseThermalConfig(),
                heat_pump=HeatPumpConfig(),
                setpoint=SetpointConfig(),
                start_day=400,
            )


class TestPhase19BisAPI:
    def test_compare_with_price_block(self, persistence) -> None:
        profile_id = _save_climate_profile(persistence, _build_model())
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/compare",
            json={
                "climate_profile_id": profile_id,
                "n_paths": 12,
                "n_years": 3,
                "seed": 0,
                "price": {"model_type": "gbm", "base_price_eur_per_kwh": 0.25,
                          "drift_annual": 0.03, "volatility_annual": 0.25},
                "house_variants": [{"label": "poor", "insulation_preset": "poor"}],
            },
        )
        assert resp.status_code == 200, resp.text
        v = resp.json()["variants"][0]
        # New fields present and the cost band is non-degenerate (price spread).
        assert "heating_kwh_annual_mean" in v and "cooling_kwh_annual_mean" in v
        assert v["annual_cost_eur_p95"] > v["annual_cost_eur_p05"]

    def test_timeseries_with_start_day(self, persistence) -> None:
        profile_id = _save_climate_profile(persistence, _build_model())
        client = _create_test_client(persistence)
        resp = client.post(
            "/api/thermal-lab/timeseries",
            json={
                "climate_profile_id": profile_id,
                "n_days": 7,
                "start_day": 181,
                "dynamic": False,
                "house": {"insulation_preset": "standard"},
            },
        )
        assert resp.status_code == 200, resp.text
        assert len(resp.json()["hours"]) == 7 * 24
